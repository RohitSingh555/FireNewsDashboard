from fastapi import APIRouter, UploadFile, File, Form, Request, Depends, HTTPException, Query
from sqlalchemy.orm import Session
from app.core.db import get_db
from app.models.excel_upload import ExcelUpload
from app.schemas.excel_upload import ExcelUploadCreate, ExcelUploadOut
import os, shutil
from datetime import datetime
from app.models.fire_news import FireNews
import openpyxl
import csv

router = APIRouter()

UPLOAD_DIR = '/app/uploads'  # Make sure this directory exists in your Docker setup
os.makedirs(UPLOAD_DIR, exist_ok=True)

def parse_datetime(dt_str):
    if not dt_str:
        return None
    for fmt in ("%Y-%m-%dT%H:%M:%S.%fZ", "%Y-%m-%dT%H:%M:%SZ", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(dt_str, fmt)
        except (ValueError, TypeError):
            continue
    return None

@router.post("/excel-uploads", response_model=ExcelUploadOut)
def upload_excel(
    request: Request,
    file: UploadFile = File(...),
    from_url: str = Form(None),
    extra: str = Form(None),
    db: Session = Depends(get_db)
):
    ip_address = request.client.host
    user_agent = request.headers.get('user-agent')
    file_name = file.filename
    # Save file to disk
    timestamp = datetime.utcnow().strftime('%Y%m%d%H%M%S')
    save_name = f"{timestamp}_{file_name}"
    file_path = os.path.join(UPLOAD_DIR, save_name)
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    print(f"Saved file: {file_path}, size: {os.path.getsize(file_path)} bytes")
    # Create DB record for upload
    upload = ExcelUpload(
        file_name=file_name,
        file_path=file_path,
        from_url=from_url,
        ip_address=ip_address,
        user_agent=user_agent,
        extra=extra
    )
    db.add(upload)
    db.commit()
    db.refresh(upload)
    # --- Parse Excel or CSV and store in fire_news ---
    inserted = 0
    if file_name.lower().endswith('.csv'):
        with open(file_path, newline='', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            for data in reader:
                if not data.get('title'):
                    continue  # Skip rows without a title
                title = data.get('title')
                published_date = parse_datetime(data.get('published_date'))
                # Check for duplicate
                exists = db.query(FireNews).filter(
                    FireNews.title == title,
                    FireNews.published_date == published_date
                ).first()
                if exists:
                    continue  # Skip duplicate
                fire_news = FireNews(
                    title=title,
                    content=data.get('content'),
                    published_date=published_date,
                    url=data.get('url'),
                    source=data.get('source'),
                    fire_related_score=data.get('fire_related_score'),
                    verification_result=data.get('verification_result'),
                    verified_at=parse_datetime(data.get('verified_at')),
                    state=data.get('state'),
                    county=data.get('county'),
                    city=data.get('city'),
                    province=data.get('province'),
                    country=data.get('country'),
                    latitude=data.get('latitude'),
                    longitude=data.get('longitude'),
                    image_url=data.get('image_url'),
                    tags=data.get('tags'),
                    reporter_name=data.get('reporter_name'),
                )
                db.add(fire_news)
                inserted += 1
        db.commit()
        print(f"Inserted {inserted} rows into fire_news from CSV")
    else:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # Always use the first row as headers, normalized to lowercase
        headers = [str(cell.value).strip().lower() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            if not data.get('title'):
                print(f"Skipping row (no title): {data}")
                continue  # Skip rows without a title
            title = data.get('title')
            published_date = parse_datetime(data.get('published_date'))
            # Check for duplicate
            exists = db.query(FireNews).filter(
                FireNews.title == title,
                FireNews.published_date == published_date
            ).first()
            if exists:
                print(f"Skipping duplicate row: {data}")
                continue  # Skip duplicate
            fire_news = FireNews(
                title=title,
                content=data.get('content'),
                published_date=published_date,
                url=data.get('url'),
                source=data.get('source'),
                fire_related_score=data.get('fire_related_score'),
                verification_result=data.get('verification_result'),
                verified_at=parse_datetime(data.get('verified_at')),
                state=data.get('state'),
                county=data.get('county'),
                city=data.get('city'),
                province=data.get('province'),
                country=data.get('country'),
                latitude=data.get('latitude'),
                longitude=data.get('longitude'),
                image_url=data.get('image_url'),
                tags=data.get('tags'),
                reporter_name=data.get('reporter_name'),
            )
            db.add(fire_news)
            inserted += 1
        db.commit()
        print(f"Inserted {inserted} rows into fire_news from Excel")
    # ---
    return upload 

@router.get("/fire-news")
def get_fire_news(
    db: Session = Depends(get_db),
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=100),
    sort_by: str = Query('published_date'),
    sort_order: str = Query('desc'),
    county: str = Query(None),
    state: str = Query(None),
    search: str = Query(None)
):
    query = db.query(FireNews)
    # Filtering
    if county:
        query = query.filter(FireNews.county == county)
    if state:
        query = query.filter(FireNews.state == state)
    # Search
    if search:
        like = f"%{search}%"
        query = query.filter((FireNews.title.ilike(like)) | (FireNews.content.ilike(like)))
    # Sorting
    sort_col = getattr(FireNews, sort_by, FireNews.published_date)
    if sort_order == 'desc':
        sort_col = sort_col.desc()
    else:
        sort_col = sort_col.asc()
    query = query.order_by(sort_col)
    # Pagination
    total = query.count()
    items = query.offset((page - 1) * page_size).limit(page_size).all()
    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "items": [
            {
                "id": n.id,
                "title": n.title,
                "content": n.content,
                "published_date": n.published_date.isoformat() if n.published_date else None,
                "url": n.url,
                "source": n.source,
                "fire_related_score": n.fire_related_score,
                "verification_result": n.verification_result,
                "verified_at": n.verified_at.isoformat() if n.verified_at else None,
                "state": n.state,
                "county": n.county,
                "city": n.city,
                "province": n.province,
                "country": n.country,
                "latitude": n.latitude,
                "longitude": n.longitude,
                "image_url": n.image_url,
                "tags": n.tags,
                "reporter_name": n.reporter_name,
                "created_at": n.created_at.isoformat() if n.created_at else None,
                "updated_at": n.updated_at.isoformat() if n.updated_at else None,
            }
            for n in items
        ]
    } 

@router.delete("/fire-news/{news_id}")
def delete_fire_news(news_id: int, db: Session = Depends(get_db)):
    news = db.query(FireNews).filter(FireNews.id == news_id).first()
    if not news:
        raise HTTPException(status_code=404, detail="Fire news entry not found")
    db.delete(news)
    db.commit()
    return {"detail": "Deleted"} 